import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

import openpyxl


LEDGER_PATH = "/root/invoice_ledger.csv"
SUMMARY_PATH = "/root/customer_summary.json"
AUDIT_PATH = "/root/correction_audit.json"
WORKBOOK_PATH = "/root/invoice_register.xlsx"
CLOSE_DATE = date(2026, 5, 31)

FIELDNAMES = [
    "row_type",
    "invoice_id",
    "original_invoice_id",
    "invoice_date",
    "due_date",
    "payment_terms_days",
    "customer_id",
    "customer_name",
    "account_owner",
    "description",
    "status",
    "currency",
    "fx_rate_to_usd",
    "amount",
    "paid_amount",
    "balance",
    "amount_usd",
    "paid_amount_usd",
    "balance_usd",
    "days_past_due",
    "aging_bucket",
    "reserved_amount_usd",
    "source_sheet",
]

SUMMARY_MONEY_KEYS = [
    "current_balance_usd",
    "days_1_30_balance_usd",
    "days_31_60_balance_usd",
    "days_61_plus_balance_usd",
    "total_amount_usd",
    "total_paid_usd",
    "total_balance_usd",
    "total_reserved_usd",
]
COUNT_KEYS = ["invoice_count", "credit_count", "open_count", "paid_count", "disputed_count", "void_count"]
FX_RATES = {"USD": Decimal("1.00"), "EUR": Decimal("1.08"), "GBP": Decimal("1.27")}
VALID_STATUSES = {"Open", "Paid", "Disputed", "Void"}
VALID_BUCKETS = {"settled", "current", "1-30", "31-60", "61+"}


def money(value):
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def ledger_rows():
    with open(LEDGER_PATH, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        assert reader.fieldnames == FIELDNAMES
        return rows


def expected_deduped_row_count():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    record_ids = set()
    for sheet_name in wb.sheetnames:
        if sheet_name in {"FX_Rates", "Customer_Aliases"}:
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        header = None
        for idx, row in enumerate(rows):
            names = {str(cell or "").strip().lower(): pos for pos, cell in enumerate(row)}
            if {"row_type", "invoice_id"}.issubset(names):
                header = (idx, names)
                break
        assert header is not None, f"No invoice header found in {sheet_name}"
        header_idx, cols = header
        for row in rows[header_idx + 1 :]:
            row_type = str(row[cols["row_type"]] or "").strip().upper()
            invoice_id = str(row[cols["invoice_id"]] or "").strip()
            if row_type in {"INVOICE", "CREDIT"} and re.fullmatch(r"(INV|CM)-\d{4}", invoice_id):
                record_ids.add(invoice_id)
    return len(record_ids)


def row_by_invoice(rows, invoice_id):
    matches = [row for row in rows if row["invoice_id"] == invoice_id]
    assert len(matches) == 1, f"Expected one row for {invoice_id}, got {len(matches)}"
    return matches[0]


def expected_aging(row):
    balance = money(row["balance_usd"])
    if balance <= Decimal("0.00"):
        return 0, "settled", Decimal("0.00")
    due_date = datetime.strptime(row["due_date"], "%Y-%m-%d").date()
    days = max(0, (CLOSE_DATE - due_date).days)
    if days == 0:
        bucket = "current"
    elif days <= 30:
        bucket = "1-30"
    elif days <= 60:
        bucket = "31-60"
    else:
        bucket = "61+"
    if row["status"] == "Disputed":
        rate = {"current": Decimal("0.25"), "1-30": Decimal("0.25"), "31-60": Decimal("0.50"), "61+": Decimal("0.75")}[bucket]
    elif row["status"] in {"Paid", "Void"} or row["row_type"] == "CREDIT":
        rate = Decimal("0.00")
    else:
        rate = {"current": Decimal("0.00"), "1-30": Decimal("0.02"), "31-60": Decimal("0.05"), "61+": Decimal("0.20")}[bucket]
    return days, bucket, (balance * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def test_ledger_schema_dedup_sort_dates_and_reference_sheets():
    rows = ledger_rows()
    assert len(rows) == expected_deduped_row_count()
    ids = [row["invoice_id"] for row in rows]
    assert len(ids) == len(set(ids)), "ledger contains duplicate invoice_id rows"
    assert [(row["invoice_date"], row["invoice_id"]) for row in rows] == sorted(
        (row["invoice_date"], row["invoice_id"]) for row in rows
    )
    for row in rows:
        datetime.strptime(row["invoice_date"], "%Y-%m-%d")
        datetime.strptime(row["due_date"], "%Y-%m-%d")
        assert row["row_type"] in {"INVOICE", "CREDIT"}
        assert row["status"] in VALID_STATUSES
        assert row["currency"] in FX_RATES
        assert money(row["fx_rate_to_usd"]) == FX_RATES[row["currency"]]
        assert row["source_sheet"].endswith("2026")


def test_fill_down_aliases_credits_and_duplicate_resolution():
    rows = ledger_rows()
    for row in rows:
        assert row["account_owner"], f"Missing account owner for {row['invoice_id']}"

    assert row_by_invoice(rows, "INV-1001")["source_sheet"] == "February 2026"
    assert row_by_invoice(rows, "INV-1005")["description"] == "Retainer corrected"
    assert row_by_invoice(rows, "INV-1013")["invoice_date"] == "2026-03-19"
    assert row_by_invoice(rows, "INV-1013")["account_owner"] == "Nora Chen"

    acme_rows = [row for row in rows if row["customer_id"] == "CUST-ALP"]
    assert acme_rows
    assert {row["customer_name"] for row in acme_rows} == {"Acme Analytics"}
    assert row_by_invoice(rows, "INV-1010")["customer_name"] == "Acme Analytics"

    cm2001 = row_by_invoice(rows, "CM-2001")
    assert cm2001["row_type"] == "CREDIT"
    assert cm2001["original_invoice_id"] == "INV-1001"
    assert money(cm2001["amount"]) < 0
    assert money(cm2001["balance_usd"]) < 0
    original_ids = {row["invoice_id"] for row in rows if row["row_type"] == "INVOICE"}
    for row in rows:
        if row["row_type"] == "CREDIT":
            assert row["original_invoice_id"] in original_ids


def test_money_fx_aging_and_status_behavior():
    rows = ledger_rows()
    money_re = re.compile(r"^-?\d+\.\d{2}$")
    statuses = {row["status"] for row in rows if row["row_type"] == "INVOICE"}
    assert {"Open", "Paid", "Disputed", "Void"}.issubset(statuses)
    assert {"USD", "EUR", "GBP"}.issubset({row["currency"] for row in rows})
    for row in rows:
        for field in ("amount", "paid_amount", "balance", "amount_usd", "paid_amount_usd", "balance_usd", "reserved_amount_usd", "fx_rate_to_usd"):
            assert money_re.fullmatch(row[field]), f"{field} is not two-decimal money: {row[field]}"
        assert money(row["balance"]) == money(row["amount"]) - money(row["paid_amount"])
        assert money(row["amount_usd"]) == (money(row["amount"]) * money(row["fx_rate_to_usd"])).quantize(Decimal("0.01"))
        assert money(row["paid_amount_usd"]) == (money(row["paid_amount"]) * money(row["fx_rate_to_usd"])).quantize(Decimal("0.01"))
        assert money(row["balance_usd"]) == (money(row["balance"]) * money(row["fx_rate_to_usd"])).quantize(Decimal("0.01"))
        days, bucket, reserve = expected_aging(row)
        assert int(row["days_past_due"]) == days
        assert row["aging_bucket"] in VALID_BUCKETS
        assert row["aging_bucket"] == bucket
        assert money(row["reserved_amount_usd"]) == reserve
        if row["status"] == "Paid" and row["row_type"] == "INVOICE":
            assert money(row["balance"]) == Decimal("0.00")
            assert money(row["paid_amount"]) == money(row["amount"])
        if row["status"] == "Void":
            assert money(row["amount"]) == Decimal("0.00")
            assert money(row["paid_amount"]) == Decimal("0.00")
            assert money(row["balance"]) == Decimal("0.00")

    inv1003 = row_by_invoice(rows, "INV-1003")
    assert inv1003["currency"] == "EUR"
    assert inv1003["aging_bucket"] == "61+"
    assert money(inv1003["reserved_amount_usd"]) == (money(inv1003["balance_usd"]) * Decimal("0.75")).quantize(Decimal("0.01"))

    inv1008 = row_by_invoice(rows, "INV-1008")
    assert inv1008["currency"] == "GBP"
    assert inv1008["status"] == "Paid"


def test_seeded_corrections_additive_payment_and_audit():
    rows = ledger_rows()
    with open(AUDIT_PATH, encoding="utf-8") as f:
        audit = json.load(f)

    assert set(audit) == {"winning_patches", "superseded", "skipped_unknown_invoice"}
    ts_values = [patch["slack_ts"] for patch in audit["winning_patches"]]
    assert ts_values == sorted(ts_values)
    for patch in audit["winning_patches"]:
        assert isinstance(patch["previous_value"], str)
        assert isinstance(patch["new_value"], str)

    inv1006 = row_by_invoice(rows, "INV-1006")
    assert inv1006["paid_amount"] == "1100.00"
    assert inv1006["balance"] == "400.00"

    inv1007 = row_by_invoice(rows, "INV-1007")
    assert inv1007["status"] == "Void"
    assert inv1007["amount"] == "0.00"

    inv1012 = row_by_invoice(rows, "INV-1012")
    assert inv1012["amount"] == "4300.00"
    assert inv1012["currency"] == "EUR"

    inv1032 = row_by_invoice(rows, "INV-1032")
    assert inv1032["amount"] == "2750.00"
    assert inv1032["paid_amount"] == "950.00"
    assert inv1032["balance"] == "1800.00"

    assert row_by_invoice(rows, "INV-1003")["description"] == "Escalated migration dispute"
    assert row_by_invoice(rows, "INV-1013")["description"] == "Store rollout phase two with change order"

    skipped = audit["skipped_unknown_invoice"]
    assert any(item["invoice_id"] == "INV-9999" and item["field"] == "paid_amount" for item in skipped)

    superseded = audit["superseded"]
    assert any(item["invoice_id"] == "INV-1006" and item["field"] == "paid_amount" for item in superseded)
    assert any(item["invoice_id"] == "INV-1012" and item["field"] == "amount" for item in superseded)
    assert any(item["invoice_id"] == "INV-1032" and item["field"] == "add_payment" for item in superseded)

    winners = {(item["invoice_id"], item["field"]): item for item in audit["winning_patches"]}
    assert winners[("INV-1006", "paid_amount")]["new_value"] == "1100.00"
    assert winners[("INV-1007", "status")]["new_value"] == "Void"
    assert winners[("INV-1012", "amount")]["new_value"] == "4300.00"
    assert winners[("INV-1032", "add_payment")]["previous_value"] == "700.00"
    assert winners[("INV-1032", "add_payment")]["new_value"] == "950.00"


def test_customer_summary_matches_final_ledger_in_usd():
    rows = ledger_rows()
    with open(SUMMARY_PATH, encoding="utf-8") as f:
        summary = json.load(f)

    assert set(summary) == {"customers", "totals"}
    customers = summary["customers"]
    assert [item["customer_id"] for item in customers] == sorted(item["customer_id"] for item in customers)

    grouped = defaultdict(list)
    for row in rows:
        grouped[row["customer_id"]].append(row)

    expected_customers = []
    expected_totals = {key: 0 for key in COUNT_KEYS}
    expected_totals.update({key: Decimal("0.00") for key in SUMMARY_MONEY_KEYS})
    for customer_id in sorted(grouped):
        group = grouped[customer_id]
        invoice_rows = [row for row in group if row["row_type"] == "INVOICE"]
        item = {
            "customer_id": customer_id,
            "customer_name": group[0]["customer_name"],
            "invoice_count": len(invoice_rows),
            "credit_count": sum(1 for row in group if row["row_type"] == "CREDIT"),
            "open_count": sum(1 for row in invoice_rows if row["status"] == "Open"),
            "paid_count": sum(1 for row in invoice_rows if row["status"] == "Paid"),
            "disputed_count": sum(1 for row in invoice_rows if row["status"] == "Disputed"),
            "void_count": sum(1 for row in invoice_rows if row["status"] == "Void"),
            "current_balance_usd": sum((money(row["balance_usd"]) for row in group if row["aging_bucket"] == "current"), Decimal("0.00")),
            "days_1_30_balance_usd": sum((money(row["balance_usd"]) for row in group if row["aging_bucket"] == "1-30"), Decimal("0.00")),
            "days_31_60_balance_usd": sum((money(row["balance_usd"]) for row in group if row["aging_bucket"] == "31-60"), Decimal("0.00")),
            "days_61_plus_balance_usd": sum((money(row["balance_usd"]) for row in group if row["aging_bucket"] == "61+"), Decimal("0.00")),
            "total_amount_usd": sum((money(row["amount_usd"]) for row in group), Decimal("0.00")),
            "total_paid_usd": sum((money(row["paid_amount_usd"]) for row in group), Decimal("0.00")),
            "total_balance_usd": sum((money(row["balance_usd"]) for row in group), Decimal("0.00")),
            "total_reserved_usd": sum((money(row["reserved_amount_usd"]) for row in group), Decimal("0.00")),
        }
        for key in COUNT_KEYS:
            expected_totals[key] += item[key]
        for key in SUMMARY_MONEY_KEYS:
            expected_totals[key] += item[key]
            item[key] = float(item[key])
        expected_customers.append(item)

    for key in SUMMARY_MONEY_KEYS:
        expected_totals[key] = float(expected_totals[key])

    assert customers == expected_customers
    assert summary["totals"] == expected_totals
    assert row_by_invoice(rows, "CM-2001")["balance_usd"] == "-150.00"
